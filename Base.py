import pandas as pd
from time import sleep
import os
import openpyxl as xl
from datetime import datetime
from UliPlot.XLSX import auto_adjust_xlsx_column_width
import win32com




class Base_ZCP015():
    def __init__(self, export_path, set_process_status, process_status) -> None:
        try:
            self.path_base = r'F:\BIOMASSA\23. Base de Dados\ZCP015.xlsx'
            self.set_msg = set_process_status
            self.set_process_status = process_status
            
            
            self.path_export = export_path
            self.rename_export_path = export_path.replace('XLSX', 'xlsx')
            os.rename(self.path_export, self.rename_export_path)
            

            self.base_file_name = self.path_base.split("\\")[-1]
            self.set_msg(f'Iniciando tratamento da base {self.base_file_name}... {self.set_process_status(11, True)}')

            # get first day of month
            #self.today = datetime.today()
            #self.date = self.today.replace(day=1, hour=0, minute=0, second=0).strftime('%Y-%m-%d %H:%M:%S')

            # get sheet names
            #self.work_base = xl.load_workbook(self.path_base) 
            #self.base_sheet=self.work_base.sheetnames[0]
            self.base_sheet = 'ZCP015'

            # read database
            self.set_msg(f'Lendo Base: {self.base_file_name}... {self.set_process_status(12, True)}')
            self.df_base = pd.read_excel(self.path_base, sheet_name=self.base_sheet)

            self.dfs = []
        except Exception as e:
            self.set_msg('Erro na inicialização!', e)
    
    
    def read_export_database(self):
        self.export_file_name = self.rename_export_path.split('/')[-1]
        # get sheet names
        self.work_book_export = xl.load_workbook(self.rename_export_path) 
        self.export_sheet=self.work_book_export.sheetnames[0]

        # read export database
        self.set_msg(f'Lendo Base SAP: {self.export_file_name}... {self.set_process_status(13, True)}')
        self.df_export = pd.read_excel(self.rename_export_path, sheet_name=self.export_sheet)
        
       


    def sort_data_pesagem(self):
        try:
            self.df_base.sort_values(by=['Dt. Pesagem Inicial', 'Hora Pesagem Inicial'], inplace=True)
            self.set_msg(f'Organizando Dt. Pesagem Inicial... {self.set_process_status(14, True)}')
        except Exception as e:
            self.set_msg('Erro ao organizar Dt. Pesagem Inicial!')

    def remove_current_values(self):
        try:
            self.set_msg(f'Removendo Linhas atuais... {self.set_process_status(15, True)}')
            self.df_base.drop(self.df_base.loc[self.df_base['Dt. Pesagem Inicial'] >= self.df_export.iloc[0]['Dt. Pesagem Inicial']].index, inplace=True)
            
        except Exception as e:
            self.set_msg('Erro ao Remover Linhas atuais!')
        
        try:
            self.set_msg(f'Removendo Linhas duplicadas... {self.set_process_status(16, True)}')
            self.new_df = self.df_base.drop_duplicates()
        except Exception as e:
            self.set_msg('Erro ao remover duplicadas!')

    def auto_adjust_column(self, df):
        try:
            self.set_msg(F'Ajustando tamanho das colunas... {self.set_process_status(20.5, True)}')
            for column in df:
                column_length = max(df[column].astype(str).map(len).max(), len(column))
                col_idx = df.columns.get_loc(column)
                self.writer.sheets[self.base_sheet].set_column(col_idx, col_idx, column_length+2)
        except Exception as e:
            self.set_msg(f'Colunas ajustadas. {self.set_process_status(20.5, True)}')
        
    

    def update_data_base(self):
        try:
            self.set_msg(f'Atualizando base: {self.base_file_name}... {self.set_process_status(17, True)}')
            self.writer = pd.ExcelWriter(self.path_base, engine='xlsxwriter', date_format='d/m/yyyy') # base file
            self.dfs.append(self.df_base)
            self.dfs.append(self.df_export)
            self.df_master = pd.concat(self.dfs, axis=False)
    
            self.set_msg(f'Base atualizada com sucesso. {self.set_process_status(17, True)}')
        except Exception as e:
            self.set_msg(f'Erro ao autualizar base {self.base_file_name}!')
    

    

    def date_format(self):
        self.set_msg(f'Ajustando formado de data... {self.set_process_status(18, True)}')
        self.df_master['Dt. Agendamento'] = pd.to_datetime(self.df_master['Dt. Agendamento'], format='%d %b %Y', errors='coerce').dt.date
        self.df_master['Dt. Pesagem Inicial'] = pd.to_datetime(self.df_master['Dt. Pesagem Inicial'], format='%d %b %Y', errors='coerce').dt.date
        
        self.df_master['Data Nota Fiscal'] = pd.to_datetime(self.df_master['Data Nota Fiscal'], format='%d %b %Y', errors='coerce').dt.date
        self.df_master['Data de criação'] = pd.to_datetime(self.df_master['Data de criação'], format='%d %b %Y', errors='coerce').dt.date
        
        self.set_msg(f'Formato de data ajustado. {self.set_process_status(18, True)}')  

    def GfConvert(self):
        self.set_msg(f'Convertendo Guia Florestal... {self.set_process_status(19, True)}')
        self.df_master['Guia Florestal'] = pd.to_numeric(self.df_master['Guia Florestal'], errors='coerce')
        self.set_msg(f"Concluido. {self.set_process_status(19, True)}")

    def save_file(self):
        try:
            self.set_msg(f"Salvando base... {self.set_process_status(20, True)}")
            self.df_master.to_excel(self.writer, sheet_name=self.base_sheet, index=False, header=True)
            self.auto_adjust_column(self.df_master)
        
            self.writer.close()
            self.set_msg(f'Salva com sucesso. {self.set_process_status(21, True)}')
            
        except Exception as e:
            self.set_msg(f'Erro ao salvar base {self.base_file_name}!')
    


    def start_update(self):
        self.read_export_database()
        self.sort_data_pesagem()
        self.remove_current_values()
        self.update_data_base()
        self.date_format()
        self.GfConvert()
        self.save_file()

        
        

